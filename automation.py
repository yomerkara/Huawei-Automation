from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils import get_column_letter 

def copy_paste_data(answers):
    source_path = r"output.xlsx"
    target_path = r"example.xlsx"

    # Load workbooks and sheets
    source_wb = load_workbook(source_path)
    target_wb = load_workbook(target_path)

    source_sheet_ip = source_wb["IP_PLAN"]
    source_sheet_1 = source_wb["3G_CDD"]
    target_sheet_transport = target_wb["Base Station Transport Data"]
    target_sheet_cell = target_wb["UMTS Cell"]

    # Columns to copy and target positions
    columns_to_copy = [
        'OAM SITE IP', 'SITE', 'SYNC SITE IP', 'IUB SITE IP', 
        'IP CLOCK PRIMARY', 'IP CLOCK SECONDARY', 
        'IUB AGG IP', 'OAM AGG IP', 'SYNC AGG IP'
    ]

    columns_to_copy_3G = [
    'Cell ID',
    'New Huawei RNC',
    'Uplink UARFCN',
    'Downlink UARFCN',
    'Max Transmit Power of Cell',
    'Cell Name',
    'DL Primary Scrambling Code(SRC)',
    'Location Area Code (LAC)',
    'CHIP',
    'PCPICH Transmit Power',
    'New Huawei RNC ID',
    'AZIMUTH']

    target_positions = {
        "transport": {
            'OAM SITE IP': [(4, "AS")],
            'SITE': [(4, "A"),(4, "E"),(4, "AN")],
            'SYNC SITE IP': [(4, "AV"),(4, "BE")],
            'IUB SITE IP': [(4, "AA"), (4, "AC")],
            'IP CLOCK PRIMARY': [(4, "AX")],
            'IP CLOCK SECONDARY': [(4, "BG")],
            'IUB AGG IP': [(4, "BO")],
            'OAM AGG IP': [(4, "BT")],
            'SYNC AGG IP': [(4, "BY")]
        },
        "cell": {
            'Cell ID': [(9, "C"), (9, "P"), (9, "Y"), (9, "AJ"), (9, "BI")],
            'New Huawei RNC': [(9, "A")],
            'Uplink UARFCN': [(9, "U"),(9, "H")],
            'Downlink UARFCN': [(9, "V"),(9, "I")],
            'Max Transmit Power of Cell': [(9, "J"),(9, "AF")],
            'Cell Name': [(9, "Q")],
            'DL Primary Scrambling Code(SRC)': [(9, "W")],
            'Location Area Code (LAC)': [(9, "AC"),(9, "X")],
            'CHIP': [(9, "AE")],
            'PCPICH Transmit Power': [(9, "AG")],
            'New Huawei RNC ID': [(9, "AW")],
            'AZIMUTH': [(9, "BE")]
        }
    }

    def get_column_index_by_name(sheet):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=2, column=col).value  # Check the value in row 2
            if cell_value == "*Product Type":
                return 2

        return None  # If "*Product Type" is not found


    # Find column indices in source sheets
    def get_column_indices(sheet, columns):
        dt = {cell.value: cell.column for cell in sheet[1] if cell.value in columns}
        print(dt)
        return dt
    
    def copy_to_next_row(sheet):
        columns = range(1, sheet.max_column + 1)
        for col in columns:
            row_3_value = sheet.cell(row=3, column=col).value
            row_4_value = sheet.cell(row=4, column=col).value
            if row_4_value is None:  # If empty, copy from row 3
                sheet.cell(row=4, column=col, value=row_3_value)

    source_ip_indices = get_column_indices(source_sheet_ip, columns_to_copy)
    source_3g_indices = get_column_indices(source_sheet_1, columns_to_copy_3G)
    # Copy data to target sheets
    def copy_data(source_sheet, column_indices, target_positions, target_sheet):
        for column_name, target_cells in target_positions.items():
            if column_name in column_indices:
                source_column = column_indices[column_name]
                for row_offset, target_col_letter in target_cells:
                    target_col = column_index_from_string(target_col_letter)
                    for row in range(2, 8):  # Adjust range as needed
                        value = source_sheet.cell(row=row, column=source_column).value
                        if column_name == 'AZIMUTH' and value is not None:
                            value *= 10
                        if column_name == 'SITE' and value is not None:
                            match answers['tech']:
                                case '900':
                                    value = 'U' + value
                                case '2100':
                                    value = 'W' + value
                        product_type_column = get_column_index_by_name(target_sheet)
                        if product_type_column is not None:
                            match answers['cabinet_type']:
                                case '3900':
                                    target_sheet.cell(row=4, column=product_type_column, value='DBS3900')
                                case '5900':
                                    target_sheet.cell(row=4, column=product_type_column, value='DBS5900')
                                case '3910':
                                    target_sheet.cell(row=4, column=product_type_column, value='DBS3910')
                                case _:
                                    print(f"Unknown cabinet type: {answers['cabinet_type']}")  # Default case if no match found
                        target_sheet.cell(row=row_offset + row - 2, column=target_col, value=value)
                        print(f"Copied '{column_name}' -> ({row}, {source_column}) to ({row_offset + row - 2}, {target_col})")

    # Copy data for transport and cell sheets
    copy_data(source_sheet_ip, source_ip_indices, target_positions["transport"], target_sheet_transport)
    copy_data(source_sheet_1, source_3g_indices, target_positions["cell"], target_sheet_cell)
    copy_to_next_row(target_sheet_transport)
    target_wb.save(target_path)
    print("Data transfer complete!")
