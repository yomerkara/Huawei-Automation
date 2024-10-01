import os
import glob
import openpyxl

def concat_sources():

    sourceDir = r"sourceFiles"

    ipSource = glob.glob(os.path.join(sourceDir, "*IP*.xlsx"))[0] 
    cddSource = glob.glob(os.path.join(sourceDir, "*CDD*.xlsx"))[0]
    output = r"./output.xlsx"

    ipFile = openpyxl.load_workbook(ipSource)
    cddFile = openpyxl.load_workbook(cddSource) #?
    ipPage =  ipFile["sRAN IP PLAN"]
    cddPage = cddFile["3G_CDD"]


    outputWorkbook = openpyxl.Workbook()
    outputSheet1 = outputWorkbook.active
    outputSheet1.title = "IP_PLAN"
    for row in ipPage.iter_rows(values_only=True):
        outputSheet1.append(row)

    outputSheet2 = outputWorkbook.create_sheet(title="3G_CDD")
    for row in cddPage.iter_rows(values_only=True):
        outputSheet2.append(row)

    outputWorkbook.save(output)

